"""Parser for sportsbookreviewsonline NCAAF historical odds Excel files (T6).

Format (per-season .xlsx, two consecutive rows per game -- visitor row then
home row; verified against tests/fixtures/flatfiles/FINDINGS.md, section 4):
columns ``Date, Rot, VH, Team, 1st, 2nd, 3rd, 4th, Final, Open, Close, ML, 2H``.
``VH`` is ``V``/``H`` for a normal game or ``N``/``N`` for a neutral-site game
(first ``N`` row is the visitor by SBR convention). Odd cell values: ``"NL"``
(no line -> None), ``"pk"``/``"PK"``/``"p"`` (pick'em -> 0.0), occasional
whitespace/case noise, and int-vs-float cell storage for what are logically
whole numbers.

Open/Close/2H disambiguation (standard SBR convention, confirmed in
FINDINGS.md): each of these columns holds *either* the point spread *or* the
game total depending on which row it's on -- there's no separate "spread"
column. For a given column, compare the visitor-row value to the home-row
value: the **larger** of the two is the game total (same on both rows in
spirit, but SBR only prints it once, on whichever row happens to hold the
larger number); the **smaller** is the point spread, and it is printed on the
favorite's row. We convert that row-relative spread to a single
home-perspective number (the sign convention used by ``betting.sbr_historical
.spread_close`` etc.): if the *home* row holds the smaller (favorite) value,
the home team is favored, so ``spread = -smaller``; if the *visitor* row holds
it, the visitor is favored and the home team is the underdog, so
``spread = +smaller``. The total is the larger value regardless of which row
it came from. The same larger=total/smaller=spread rule applies independently
to Open, Close, and 2H (2H = second-half line).

Required behavior:
- Pair rows strictly (V then H, or two N rows for neutral); raise
  ``ParserStructureError`` with row context when pairing breaks.
- Season inference: ``ctx.season`` if set, else the first 4-digit number in
  ``ctx.file_name`` (e.g. "ncaa football 2013-14.xlsx" -> 2013); game dates
  Aug-Dec belong to the season year, Jan dates to season year + 1.
- Yield one row per game matching ``betting.sbr_historical`` columns, with
  ``home_team``/``away_team`` carrying the SOURCE spellings (the framework
  resolves via crosswalk and preserves originals to ``*_source`` --
  ``keep_source_names=True`` on the registry spec).
- Normalize "NL" -> None, "pk"/"PK"/"p" -> 0.0; coerce numerics defensively;
  fail loud (``ParserStructureError``) on per-year format drift (bad header,
  broken pairing, non-numeric junk that isn't NL/pk) rather than guessing.
"""

import io
import re
from collections.abc import Iterator
from datetime import date

from openpyxl import load_workbook

from ..flat_files import ParseContext, ParserStructureError

# Confirmed column set per FINDINGS.md section 4 (live HTML table, 2022-23 season).
EXPECTED_COLUMNS = frozenset(
    {"Date", "Rot", "VH", "Team", "1st", "2nd", "3rd", "4th", "Final", "Open", "Close", "ML", "2H"}
)

_PICKEM_TOKENS = frozenset({"PK", "P"})
_NO_LINE_TOKEN = "NL"


def parse(raw: bytes, ctx: ParseContext) -> Iterator[dict]:
    """Parse an SBR season Excel file into betting.sbr_historical rows."""
    wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    ws = wb.worksheets[0]
    rows_iter = ws.iter_rows(values_only=True)

    try:
        header = next(rows_iter)
    except StopIteration as e:
        raise ParserStructureError(f"sbr: {ctx.file_name!r} is empty (no header row)") from e

    col_index = _build_column_index(header)
    actual = set(col_index)
    missing = EXPECTED_COLUMNS - actual
    extra = actual - EXPECTED_COLUMNS
    if missing or extra:
        raise ParserStructureError(
            f"sbr: {ctx.file_name!r} header mismatch -- "
            f"missing={sorted(missing)}, extra={sorted(extra)}"
        )

    season = ctx.season if ctx.season is not None else _season_from_filename(ctx.file_name)
    if season is None:
        raise ParserStructureError(
            f"sbr: no season available -- ctx.season is unset and no 4-digit year found in "
            f"file_name {ctx.file_name!r}"
        )

    data_rows = [
        (row_num, row)
        for row_num, row in enumerate(rows_iter, start=2)
        if row is not None and any(v is not None for v in row)
    ]

    if len(data_rows) % 2 != 0:
        last_num, last_row = data_rows[-1]
        raise ParserStructureError(
            f"sbr: unpaired trailing row {last_num} (odd row count): {last_row!r}"
        )

    def cell(row: tuple, name: str):
        idx = col_index[name]
        return row[idx] if idx < len(row) else None

    for i in range(0, len(data_rows), 2):
        num1, row1 = data_rows[i]
        num2, row2 = data_rows[i + 1]
        vh1 = str(cell(row1, "VH") or "").strip().upper()
        vh2 = str(cell(row2, "VH") or "").strip().upper()

        if vh1 == "V" and vh2 == "H":
            visitor_num, visitor_row = num1, row1
            home_num, home_row = num2, row2
            neutral_site = False
        elif vh1 == "N" and vh2 == "N":
            # First N row is the visitor by SBR convention (confirmed in FINDINGS.md).
            visitor_num, visitor_row = num1, row1
            home_num, home_row = num2, row2
            neutral_site = True
        else:
            raise ParserStructureError(
                f"sbr: bad row pairing at rows {num1}-{num2} -- VH={vh1!r}/{vh2!r} "
                f"row{num1}={row1!r} row{num2}={row2!r}"
            )

        game_date = _parse_game_date(cell(visitor_row, "Date"), season, visitor_num)

        away_team = str(cell(visitor_row, "Team")).strip()
        home_team = str(cell(home_row, "Team")).strip()

        away_rot = _to_int(cell(visitor_row, "Rot"), visitor_num, "Rot")
        home_rot = _to_int(cell(home_row, "Rot"), home_num, "Rot")

        away_final = _to_optional_int(cell(visitor_row, "Final"), visitor_num, "Final")
        home_final = _to_optional_int(cell(home_row, "Final"), home_num, "Final")

        spread_open, total_open = _spread_and_total(
            cell(visitor_row, "Open"), cell(home_row, "Open"), visitor_num, home_num, "Open"
        )
        spread_close, total_close = _spread_and_total(
            cell(visitor_row, "Close"), cell(home_row, "Close"), visitor_num, home_num, "Close"
        )
        spread_2h, total_2h = _spread_and_total(
            cell(visitor_row, "2H"), cell(home_row, "2H"), visitor_num, home_num, "2H"
        )

        away_ml = _to_optional_ml(cell(visitor_row, "ML"), visitor_num, "ML")
        home_ml = _to_optional_ml(cell(home_row, "ML"), home_num, "ML")

        yield {
            "season": season,
            "game_date": game_date,
            "home_team": home_team,
            "away_team": away_team,
            "home_rot": home_rot,
            "away_rot": away_rot,
            "home_final": home_final,
            "away_final": away_final,
            "spread_open": spread_open,
            "spread_close": spread_close,
            "total_open": total_open,
            "total_close": total_close,
            "home_ml": home_ml,
            "away_ml": away_ml,
            "spread_2h": spread_2h,
            "total_2h": total_2h,
            "neutral_site": neutral_site,
        }


def _build_column_index(header: tuple) -> dict[str, int]:
    """Map header cell name -> column index, skipping blank/None cells."""
    col_index: dict[str, int] = {}
    for idx, raw_name in enumerate(header):
        if raw_name is None:
            continue
        name = str(raw_name).strip()
        if name and name not in col_index:
            col_index[name] = idx
    return col_index


def _season_from_filename(file_name: str | None) -> int | None:
    """First 4-digit number in the filename, e.g. 'ncaa football 2013-14.xlsx' -> 2013."""
    if not file_name:
        return None
    m = re.search(r"\d{4}", file_name)
    return int(m.group()) if m else None


def _parse_game_date(raw_date, season: int, row_num: int) -> date:
    """Date cells are MMDD ints; Aug-Dec -> season year, Jan -> season year + 1."""
    date_int = _to_int(raw_date, row_num, "Date")
    month, day = divmod(date_int, 100)
    if month >= 8:
        year = season
    elif month <= 1:
        year = season + 1
    else:
        raise ParserStructureError(
            f"sbr: row {row_num}: Date {raw_date!r} decodes to month {month}, outside the "
            "expected Aug-Jan CFB season range"
        )
    try:
        return date(year, month, day)
    except ValueError as e:
        raise ParserStructureError(
            f"sbr: row {row_num}: Date {raw_date!r} -> invalid calendar date "
            f"{year}-{month}-{day}: {e}"
        ) from e


def _to_int(raw, row_num: int, label: str) -> int:
    """Strict whole-number coercion (Date, Rot) -- no NL/pk sentinels expected here."""
    if isinstance(raw, bool) or raw is None:
        raise ParserStructureError(f"sbr: row {row_num}: {label} value {raw!r} is not numeric")
    if isinstance(raw, int):
        return raw
    if isinstance(raw, float):
        if raw.is_integer():
            return int(raw)
        raise ParserStructureError(
            f"sbr: row {row_num}: {label} value {raw!r} is not a whole number"
        )
    text = str(raw).strip()
    try:
        return int(text)
    except ValueError:
        pass
    try:
        as_float = float(text)
    except ValueError as e:
        raise ParserStructureError(
            f"sbr: row {row_num}: {label} value {raw!r} is not numeric"
        ) from e
    if not as_float.is_integer():
        raise ParserStructureError(
            f"sbr: row {row_num}: {label} value {raw!r} is not a whole number"
        )
    return int(as_float)


def _to_optional_int(raw, row_num: int, label: str) -> int | None:
    """Whole-number coercion allowing 'NL' -> None (e.g. postponed/no-line Final)."""
    if raw is None:
        return None
    if isinstance(raw, str) and raw.strip().upper() == _NO_LINE_TOKEN:
        return None
    return _to_int(raw, row_num, label)


def _to_optional_ml(raw, row_num: int, label: str) -> int | None:
    """Moneyline coercion: 'NL' -> None, else int; no pk sentinel expected for ML."""
    if raw is None:
        return None
    if isinstance(raw, bool):
        raise ParserStructureError(f"sbr: row {row_num}: {label} value {raw!r} is not numeric")
    if isinstance(raw, (int, float)):
        return _to_int(raw, row_num, label)
    text = str(raw).strip()
    if text == "":
        return None
    if text.upper() == _NO_LINE_TOKEN:
        return None
    try:
        as_float = float(text)
    except ValueError as e:
        raise ParserStructureError(
            f"sbr: row {row_num}: {label} value {raw!r} is not numeric or NL"
        ) from e
    return int(as_float)


def _to_line(raw, row_num: int, label: str) -> float | None:
    """Open/Close/2H cell coercion: 'NL' -> None, 'pk'/'PK'/'p' -> 0.0, else float."""
    if raw is None:
        return None
    if isinstance(raw, bool):
        raise ParserStructureError(f"sbr: row {row_num}: {label} value {raw!r} is not numeric")
    if isinstance(raw, (int, float)):
        return float(raw)
    text = str(raw).strip()
    if text == "":
        return None
    upper = text.upper()
    if upper == _NO_LINE_TOKEN:
        return None
    if upper in _PICKEM_TOKENS:
        return 0.0
    try:
        return float(text)
    except ValueError as e:
        raise ParserStructureError(
            f"sbr: row {row_num}: {label} value {raw!r} is not numeric, NL, or pk"
        ) from e


def _spread_and_total(
    visitor_raw, home_raw, visitor_num: int, home_num: int, label: str
) -> tuple[float | None, float | None]:
    """Disambiguate one Open/Close/2H column pair into (home-perspective spread, total).

    See module docstring for the larger=total/smaller=spread convention. Both
    cells must agree on whether a line exists: if one side is 'NL' and the
    other has a real value, that is treated as format drift (fail loud rather
    than guess which value the missing side would have taken).
    """
    v = _to_line(visitor_raw, visitor_num, f"{label} (visitor)")
    h = _to_line(home_raw, home_num, f"{label} (home)")

    if v is None and h is None:
        return None, None
    if v is None or h is None:
        raise ParserStructureError(
            f"sbr: rows {visitor_num}-{home_num}: {label} has NL on only one side "
            f"(visitor={visitor_raw!r}, home={home_raw!r})"
        )

    total = max(v, h)
    smaller = min(v, h)
    if v < h:
        spread = smaller  # visitor favored -> home is the underdog -> positive home spread
    elif h < v:
        spread = -smaller  # home favored -> negative home spread
    else:
        spread = 0.0
    return spread, total
